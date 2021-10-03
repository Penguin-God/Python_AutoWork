import datetime
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# formula : 공식, 공식 규격
ws.column_dimensions["A"].width = 25
ws["A1"] = datetime.datetime.today() # 오늘 날짜, 시간 정보
ws["A2"] = "=SUM(1, 5, 22)" # 합
ws["A3"] = "=AVERAGE(2,5,11)" # 평균

ws["A4"] = 30
ws["A5"] = 30
ws["A6"] = 30
ws["A7"] = "=SUM(A4:A6)"

wb.save("C:/Users/parkj/Desktop/프로그래밍/python/RPA/Excel/File/Sample_Formula.xlsx")