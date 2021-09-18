# openpyxl 설치
from openpyxl import Workbook
wb = Workbook() # 새로운 Workbook 생성
# sheet는 엑셀 파일 열면 아래에 있는 탭 비슷한 거
# wb.active 현재 활성화된 sheet 가져옴 (현재는 wb)
ws = wb.create_sheet("Hello Wolrd", 0)  # 액셀에 주어진 이름으로 주어진 index에 새로운 시트 생성
ws.sheet_properties.tabColor = "00ccff"
ws["A1"] = "Hello A1" # 엑셀 A1 셀에 값 대입
copySheet = wb.copy_worksheet(ws) # 시트 복사
copySheet.title = "Good Copy" 

print(wb.sheetnames) # 모든 시트 이름 확인

wb.save("Sample.xlsx")
wb.close()


