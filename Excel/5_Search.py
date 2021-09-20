from openpyxl import load_workbook
wb = load_workbook("C:/Users/parkj/Desktop/프로그래밍/python/RPA/Excel/File/Sample.xlsx")
ws = wb.active

# 영어점수 80점 이상인 애들 찾기
for row in ws.iter_rows(min_row=2):
    if(int(row[2].value) >= 80):
        print(row[0].value, ":", row[2].value)

for row in ws.iter_rows(max_row=1):
    for cell in row:
        if(cell.value == "영어"):
            cell.value = "정보"

# modifide : 수정하다, 바꾸다
wb.save("C:/Users/parkj/Desktop/프로그래밍/python/RPA/Excel/File/Sample_Modifide.xlsx")