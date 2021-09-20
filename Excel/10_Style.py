from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
wb = load_workbook("C:/Users/parkj/Desktop/프로그래밍/python/RPA/Excel/File/Sample.xlsx")
ws = wb.active

a1 = ws["A1"] # 번호
b1 = ws["B1"] # 수학
c1 = ws["C1"] # 영어

# dimension : 치수, 차원, 크기
# A열의 넓이 10으로 설정
ws.column_dimensions["A"].width = 10
# 1행의 높이 20으로 설정
ws.row_dimensions[1].height = 30

# a1폰트 빨간색, 이탤릭체(글자 눕히는거), bold 적용
a1.font = Font(color="FF0000", italic=True, bold=True)
# name : 폰트명, strike : 나무위키 몸통 밑줄
b1.font = Font(color="CC33FF", name="Arial", strike=True)
# 사이즈 20, 밑줄 적용
c1.font = Font(color="0000FF", size=20, underline="single")


# 테두리 적용
defaultBorder = Border(right=Side(style="medium"), left=Side(style="medium"), top=Side(style="medium"), bottom=Side(style="medium"))

a1.border = defaultBorder 
b1.border = defaultBorder 
c1.border = defaultBorder

# 모든 셀 중앙 정렬
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")

# 80점 넘는 셀에 대해서 초록색으로 적용
for row in ws.iter_rows(min_col=2, min_row=2):
    for cell in row:
        if(int(cell.value) >= 80):
            # 배경 색깔 정의
            cell.fill = PatternFill(fgColor="0cb9e4", fill_type="solid")
            cell.font = Font(color="FF0000")
            cell.border = defaultBorder

# isinstance(cell.value, int) : 인트형이면 true

# 틀 고정 : 지정한 셀의 위쪽과 왼쪽이 고정됨
ws.freeze_panes = "B2"


wb.save("C:/Users/parkj/Desktop/프로그래밍/python/RPA/Excel/File/Sample_Style.xlsx")