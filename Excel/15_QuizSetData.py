'''
나는 소프트웨어학과 조교(노예)로서 학생들의 등급을 매겨야 함
과목 점수 비중은 아래와 같음
출석 - 10
퀴즈1 - 10
퀴즈2 - 10
중간고사 - 20
기말고사 - 30
프로젝트 - 20

1. 퀴즈2번은 문제가 있어 전원 만점처리
2. H열에 총점(=SUM 사용) I열에 등급 정보 추가
3. 총점 90이상 A, 80 이상 B, 70 이상 C, 나머지D
4. 출석 5 미만은 무조건 F

학번, 출석, 퀴즈1, 퀴즈2, 중간고사, 기말고사, 프로젝트
1,10,8,5,14,26,12
2,7,3,7,15,24,18
3,9,5,8,8,12,4
4,7,8,7,17,21,18
5,7,8,7,16,25,15
6,3,5,8,8,17,0
7,4,9,10,16,27,18
8,6,6,6,15,19,17
9,10,10,9,19,30,19
10,9,8,8,20,25,20
'''

from openpyxl import load_workbook
wb = load_workbook("C:/Users/parkj/Desktop/프로그래밍/python/RPA/Excel/File/QuizData.xlsx")
ws = wb.active

#데이터 세팅
'''
나도코딩의 깔끔한 데이터 세팅 (나는 기존 엑셀 데이터를 덮어쓰는 방식으로 구현해서 apeend는 못씀)

ws.append(("학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트", "총점", "등급"))
data = [ (1,10,8,5,14,26,12),
(2,7,3,7,15,24,18),
(3,9,5,8,8,12,4),
(4,7,8,7,17,21,18),
(5,7,8,7,16,25,15),
(6,3,5,8,8,17,0),
(7,4,9,10,16,27,18),
(8,6,6,6,15,19,17),
(9,10,10,9,19,30,19),
(10,9,8,8,20,25,20) ]

for row in data:
    ws.append(row)
'''

# 내가 한 데이터 세팅

# 타이틀
titleData = ["학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트", "총점", "등급"]
for title in range(0, len(titleData)):
    ws.cell(column=title+1, row = 1 , value = titleData[title]) 

# 데이터
data = [ [1,10,8,5,14,26,12],
[2,7,3,7,15,24,18],
[3,9,5,8,8,12,4],
[4,7,8,7,17,21,18],
[5,7,8,7,16,25,15],
[6,3,5,8,8,17,0],
[7,4,9,10,16,27,18],
[8,6,6,6,15,19,17],
[9,10,10,9,19,30,19],
[10,9,8,8,20,25,20] ]

for x in range(1, len(data[0]) + 1):
    for y in range(1, len(data) + 1):
        # row는 2번째줄부터
        ws.cell( column=x, row= 1 + y, value=data[y - 1][x - 1])

'''
박준의 비효율적인 1,2번 문제 풀이 (ws["A"] 이렇게 접근할 수 있는줄 몰랐고, enumerate를 몰랐음)

#퀴즈 2번 만점
for i in range(2, ws.max_row+1):
    ws.cell(row=i, column=4, value=10)

#=SUM으로 점수 총합 구함
scoreCount = 1
for row in ws.iter_rows(min_row=2, min_col=2, max_col=7):
    formulaData = row[0].coordinate + ":" + row[len(row)-1].coordinate
    scoreCount += 1
    ws.cell(column=8, row=scoreCount, value="=SUM"+ "(" + formulaData + ")")
'''

# #퀴즈 2번 만점
# for i, cell in enumerate(ws["D"]):
#     if(i == 0) : continue
#     # print(cell) # 값이 있는 D colunm의 cell객체 정보 가져옴
#     cell.value = 10

# # =SUM으로 점수 총합 구함
# for i, cell in enumerate(ws["H"], start=1):
#     if(i == 1) : continue
#     cell.value = "=SUM(B{}:G{})".format(i, i)


def ReturnGrade(score):
    if(score >= 90): return "A"
    elif(score >= 80): return "B"
    elif(score >= 70): return "C"
    else: return "D"

# 총합을 따로 구해서 여기서 한번에 등급까지 할 수도 있음
def SetFile():
    for i, scores in enumerate(data, start=2):
        # 총점
        ws.cell(row=i, column=8).value = "=SUM(B{}:G{})".format(i, i)
        # 등급
        ws.cell(row=i, column=9).value = ReturnGrade(sum(scores[1:]) - scores[3] + 10 )
        # 출석 점수 5 미만이면 F
        if(int(ws.cell(row=i, column=2).value) < 5) : ws.cell(row=i, column=9, value="E")

SetFile()

wb.save("C:/Users/parkj/Desktop/프로그래밍/python/RPA/Excel/File/QuizData.xlsx")