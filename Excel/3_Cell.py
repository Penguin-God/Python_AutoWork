from typing import Counter
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws["A1"] = "이것은 A1이다"
ws["B1"].value = "이것은 B1이다"

print(ws["A1"]) # sheet의 객체 정보 출력
print(ws["A1"].value) # A1의 값 출력

# 0이 아닌 1부터 시작
# row = 1, 2, 3...
# column = A(1), B(2), C(3)... 
ws.cell(column = 2, row = 2).value = "이것은 2 X 2 이다" # cell은 값 바꾸려면 .value붙여야됨
print(ws.cell(column=2, row=2).value)
print(ws["A13"].value) # 값이 없을때는 None출력

c2 = ws.cell(column=3, row=2, value="This is C2")
c2 = "Hello C" # 셀값에 적용 안됨 (주소가 객체처럼 작동 안하는듯)

index = 1
# row와 column중 어느 순서로 값이 대입되는지 반복문 순서 주의 (지금은 컬럼부터 커짐)
for x in range(1, 11): # row 10줄
    for y in range(1, 11): # column 10줄
        ws.cell(row=x, column=y, value=index) 
        index += 1

wb.save("C:/Users/parkj/Desktop/프로그래밍/python/RPA/Excel/File/Sample.xlsx")