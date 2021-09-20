from openpyxl import load_workbook
from random import *
wb = load_workbook("C:/Users/parkj/Desktop/프로그래밍/python/RPA/Excel/File/Sample.xlsx")
ws = wb.active

# 번호 수학 영어 -> 번호 국어 영어 수학 으로 바꾸기
#ws.move_range("B1:C11") # B1 ~ C11 까지 드래그한거랑 비슷함 (이동 하려는 범위 정의)
# row 0줄 column2줄 옮김 -로 하면 뒤로감
ws.move_range("B1:B11", rows=0, cols=2) # 옮기는 위치에 value가 존재하면 덮어씀
ws["B1"].value = "국어"

for i in range(2, 12):
    ws.cell(row=i, column=2, value=randint(0,100))

wb.save("C:/Users/parkj/Desktop/프로그래밍/python/RPA/Excel/File/Sample_Move.xlsx")