from openpyxl import load_workbook
wb = load_workbook("Sample.xlsx")
ws = wb.active

# max_row는 값이 있는 최대로우의 값인데 range는 1작게 나오니까 +1 함 
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x,column=y).value, end = "  ")
    print("\n")