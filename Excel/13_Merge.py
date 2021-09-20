from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 셀 병합
ws.merge_cells("B1:D4") # 병합할 셀 범위 정의
ws["B1"] = "격백"

ws.merge_cells("E1:F4")
ws["E1"] = "억울"
ws.unmerge_cells("E1:F4") # 병합 취소 (범위 안맞으면 에러 뜸)

wb.save("C:/Users/parkj/Desktop/프로그래밍/python/RPA/Excel/File/Sample_Merge.xlsx")