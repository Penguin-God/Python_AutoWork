from openpyxl import Workbook
from random import *
from openpyxl.utils.cell import coordinate_from_string
wb = Workbook()
ws = wb.active

# append하면 한줄에 순서대로 대입 가능 한 줄에 넣는 거기 때문에 인자도 리스트로 받음
ws.append(["번호", "수학", "영어"])
for i in range(1, 11):
    ws.append([i, randint(0,100), randint(0,100)])

column_B = ws["C"] # C컬럼의 데이터의 객체 정보를 가져옴(값이 있는거만)

# B부터 C컬럼까지의 데이터를 리스트 형식으로 가져옴 (한 줄당 리스트에 데이터 하나씩 추가)
column_Score = ws["B:C"]
# column_Score가 가지고 있는 B,C 컬럼의 데이터 출력
# for column in column_Score:
#     for score in column:
#         print(score.value)

rowTitle = ws[1] # row1번째 줄
rowScore = ws[2:6] # row 2~6번째 줄 데이터 가져옴
anyRowScore = ws[2:ws.max_row] # row 2 ~ 마지막 줄 까지의 데이터 (값 있는 줄까지)
# for row in rowScore:
#     for score in row:
#         #print(score.value, end=" ")
#         # coordinate : 동등한, 좌표의, 조정하다
#         #print(score.coordinate, end=" ") # 데이터의 셀 위치 출력

#         xy = coordinate_from_string(score.coordinate)
#         #print(xy) # 튜플에 담아서 좌표처럼 사용가능
#         print(xy[0], end=", ") # column
#         print(xy[1]) # row
#     print('\n')

# 전체 row
#print(tuple(ws.rows))
#print(ws.iter_rows())
# 전체 column
#print(tuple(ws.columns))
#print(ws.iter_cols())

# column B,C의 1~ 9번째 줄의 데이터 가져옴 (인자 기본값은 첫번째줄 ~ 마지막줄)
for row in ws.iter_rows(max_row=9, min_col=2, max_col=4):
    print(row[0].value, row[1].value) # 가져온 row 중 1,2번째 인덱스의 값(B,C) 출력

wb.save("C:/Users/parkj/Desktop/프로그래밍/python/RPA/Excel/File/Sample.xlsx")
