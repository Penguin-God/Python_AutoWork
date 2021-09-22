'''
나는 소프트웨어학과 조교(노예)로 등급을 매겨야 함
과목 점수 비중은 아래와 같음
출석 - 10
퀴즈1 - 10
퀴즈2 - 10
중간고사 - 20
기말고사 - 30
프로젝트 - 20

1. 퀴즈2번은 문제가 있어 전원 만점처리
2. H열에 총점(=SUM 사용) I열에 등급 정보 추가
3. 총점 90이상 A, 80 이상 B, 70 이상 C, 나머지D
4. 출석 5 미만은 무조건 F
'''

from openpyxl import load_workbook
wb = load_workbook("C:/Users/parkj/Desktop/프로그래밍/python/RPA/Excel/File/Quiz.xlsx", data_only=True)
ws = wb.active

for i in range(2, ws.max_row + 1):
    ws.cell(column=4, row=i, value=10)


    