from openpyxl import load_workbook
wb = load_workbook("C:/Users/parkj/Desktop/프로그래밍/python/RPA/Excel/File/QuizData.xlsx", data_only=True)
ws = wb.active

def ReturnGrade(score):
    if(score >= 90): return "A"
    elif(score >= 80): return "B"
    elif(score >= 70): return "C"
    else: return "D"

# 점수에 따른 등급 매김
for i, cell in enumerate(ws["I"], start=1):
    if(i == 1) : continue
    cell.value = ReturnGrade(int(ws.cell(row=i, column=8).value))

# 출석 5미만이면 F
for i, cell in enumerate(ws["B"], start=1):
    if(i == 1) : continue
    if(int(cell.value) < 5):
        ws.cell(row=i, column=9, value="E")

wb.save("C:/Users/parkj/Desktop/프로그래밍/python/RPA/Excel/File/QuizAnswer.xlsx")